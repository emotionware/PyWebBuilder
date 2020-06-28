

import sys, os, codecs, re
import json
import urllib
import urllib.request
import smtplib
import csv
import io
import flask_excel
import time


from xhtml2pdf import pisa

from random import randint,uniform,random


from urllib.parse import unquote
 

from email.message import EmailMessage
from flask_mail import Mail, Message
from datetime import datetime
from flask import Flask,  jsonify,request, render_template, session,redirect, make_response
from flask import send_from_directory
import flask_excel as excel

from email.utils import make_msgid
import mimetypes

from openpyxl import Workbook 
from openpyxl import load_workbook


from os import listdir
from os.path import isfile, join


from werkzeug.utils import secure_filename

import os.path
from os import path
 

from _4TPyModel import app
from _4TPyModel.code.wservice import libros as libreria
from _4TPyModel.code.navbar import navbartemplate as navbartemplate
from _4TPyModel.code.topmenu import topmenutemplate 
from _4TPyModel.code.topmenu import topmenutemplate2
from _4TPyModel.code.gridpopulate import populatearray
from _4TPyModel.code.gridpopulate import populate
from _4TPyModel.code.gridpopulate import url_for
from _4TPyModel.code.gridpopulate import populatepropiedad
from _4TPyModel.code.gridpopulate import populatexlsx
from _4TPyModel.code.gridpopulate import valtoken
from _4TPyModel.code.gridpopulate import populatequery
from _4TPyModel.code.gridpopulate import populatequery2
from _4TPyModel.code.dbexec import dbexecutor
from _4TPyModel.code.gridpopulate import buscador
from _4TPyModel.code.gridpopulate import buscadorxlsx
from _4TPyModel.code.gridpopulate import quierolaclasusulawhere
from _4TPyModel.code.gridpopulate import cuentaregistros
from _4TPyModel.code.gridpopulate import buscadorarray
from _4TPyModel.code.gridpopulate import resolversentencia
from _4TPyModel.code.gridpopulate import quieroelid
from _4TPyModel.code.gridpopulate import dimetipo
from _4TPyModel.code.gridpopulate import checarpermisos




#Project Paths
from _4TPyModel.code.settings import APP_STATIC
from _4TPyModel.code.settings import APP_TEMPLATE 
from _4TPyModel.code.settings import APP_ROOT
from _4TPyModel.code.settings import APP_CODE
from _4TPyModel.code.settings import APP_UPLOADS


#String Connection
from _4TPyModel.code.dbstring import hostx
from _4TPyModel.code.dbstring import userx
from _4TPyModel.code.dbstring import passwdx
from _4TPyModel.code.dbstring import databasex
from _4TPyModel.code.dbstring import portx
from _4TPyModel.code.dbstring import GMTZone

#SMTP
from _4TPyModel.code.smtp import MAIL_SERVERx
from _4TPyModel.code.smtp import MAIL_PORTx
from _4TPyModel.code.smtp import MAIL_USE_TLSx
from _4TPyModel.code.smtp import MAIL_USE_SSLx
from _4TPyModel.code.smtp import MAIL_DEFAULT_SENDERx
from _4TPyModel.code.smtp import MAIL_USERNAMEx
from _4TPyModel.code.smtp import MAIL_PASSWORDx
from _4TPyModel.code.smtp import MAIL_DEBUGx
from _4TPyModel.code.smtp import MAIL_SUPPRESS_SENDx
from _4TPyModel.code.smtp import TESTINGx



#Builders
from _4TPyModel.code.builderconfigurator import builderconfigurator
from _4TPyModel.code.buildergrid import buildergrid
from _4TPyModel.code.builderform import builderform
from _4TPyModel.code.buildergoogleform import builderform as googlebuilder
from _4TPyModel.code.builderadaptation import builderadaptation



from _4TPyModel.code.gridpopulate import getfields
from _4TPyModel.code.gridpopulate import getfields2
from _4TPyModel.code.gridpopulate import damedatosobjeto
from _4TPyModel.code.gridpopulate import damedisplay
from _4TPyModel.code.gridpopulate import dameultimoid
from _4TPyModel.code.gridpopulate import tablasparabuild
from _4TPyModel.code.gridpopulate import tablasparabuildscript
from _4TPyModel.code.gridpopulate import damedestinatariosdealta
from _4TPyModel.code.gridpopulate import volcadovista
from _4TPyModel.code.gridpopulate import damedestinatariosdecambioestatus
from _4TPyModel.code.gridpopulate import esvalido
from _4TPyModel.code.gridpopulate import checaradmin
from _4TPyModel.code.gridpopulate import checaremail

from _4TPyModel.code.gridpopulate import cacheregen
from _4TPyModel.code.gridpopulate import purgar
from _4TPyModel.code.gridpopulate import agendadatos
from _4TPyModel.code.gridpopulate import generarelementos


# Make the WSGI interface available at the top level so wfastcgi can get it.
wsgi_app = app.wsgi_app

#show errors
app.debug = True

app.config['ENV'] = 'development'
app.config['DEBUG'] = True
app.config['TESTING'] = True

excel.init_excel(app)

 
app.config['MAIL_SERVER']=MAIL_SERVERx
app.config['MAIL_PORT']=MAIL_PORTx
app.config['MAIL_USE_TLS']= MAIL_USE_TLSx
app.config['MAIL_USE_SSL']= MAIL_USE_SSLx
app.config['MAIL_DEFAULT_SENDER'] = MAIL_DEFAULT_SENDERx
app.config['MAIL_USERNAME']=MAIL_USERNAMEx
app.config['MAIL_PASSWORD']=MAIL_PASSWORDx
app.config['MAIL_DEBUG']=MAIL_DEBUGx
app.config['MAIL_SUPPRESS_SEND ']= MAIL_SUPPRESS_SENDx
app.config['TESTING']= TESTINGx
mail = Mail(app)


#Required for manage Sessions Var
app.secret_key = 'You Will Never Guess'




#upload settings
UPLOAD_FOLDER = APP_STATIC
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 80 * 1024 * 1024
ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg','ppt','pptx','odp', 'gif', 'xls', 'xlsx', 'doc', 'docx'])


def productividad(accion,idusername,username,tabla):
    
    laorden="insert into tblproductividad (accion,idsysuser,username,tabla,fecha,hora,fechaalta) values('" + str(accion) + "','" + str(idusername) + "','" + str(username) + "','" + str(tabla) + "',date(now()),time(now()),now());"
    
    print(laorden)

    res=dbexecutor.executor(laorden)

@app.route('/pivotfield')
def pivotfield():
    return render_template('config/pivotfield.js')

@app.route('/diagramitems')
def diagramitems():
    return render_template('config/diagramitems.js')

@app.route('/host')
def host():
    return render_template('host.html',topmenu=topmenutemplate())


@app.route('/')
def start():
    #return render_template('config/home.html',topmenu=topmenutemplate())
    return redirect ("/host")

@app.route('/render', methods=['GET', 'POST'])
def render():
    return render_template('tables/' + request.args.get('archivo')  + '.html',topmenu=topmenutemplate())

@app.route('/mailclient')
def mailclient():
    return render_template('mailclient.html',topmenu=topmenutemplate())

@app.route('/json2csv', methods=['GET', 'POST'])
def json2csv():

    data =[]
    buscar = request.args.get('buscar')
    tablename = request.args.get('tablename')
    CAMPODISPLAY_startsWith = request.args.get('CAMPODISPLAY_startsWith')
    campofiltro = request.args.get('campofiltro')
    idobjetofiltro = request.args.get('idobjetofiltro')

    if buscar=='' or buscar=='null' or buscar==None:
        data=populatearray(tablename,CAMPODISPLAY_startsWith,campofiltro,idobjetofiltro)
        return excel.make_response(data, "csv")
    else:
        data=buscadorarray("select *,concat('<a href=""javascript:modoedicion(',ID" + tablename.replace('tbl','') + ",');"">Editar</a>') as EDITAR from " + tablename.lower() + " where " + quierolaclasusulawhere(tablename,buscar))
        return excel.make_response_from_array(data, "csv")





@app.route('/ayuda')
def ayuda():
    return render_template('config/help.html',topmenu=topmenutemplate())


@app.route('/home')
def inicio():

    return render_template('config/home.html',topmenu=topmenutemplate() )

    #sentencia="update morgan"
    #return ejecutor.exec(sentencia)


@app.route('/mobile')
def mobile():

    return render_template('config/mobile.html',topmenu2=topmenutemplate2() )

    #sentencia="update morgan"
    #return ejecutor.exec(sentencia)

@app.route('/enviaremail', methods=['GET', 'POST'])
def enviaremail():
    subject = request.args.get('subject')
    destinatario = request.args.get('destinatario')
    bodytext = request.args.get('bodytext')

    folder1 = request.args.get('folder1')
    folder2 = request.args.get('folder2')
    elnombre = request.args.get('elnombre')


    msg = Message(subject,
              recipients=[destinatario])
    msg.recipients = [destinatario]
    msg.body = bodytext
    msg.html = bodytext


    if elnombre!=None:
        pathx=os.path.join(app.root_path, folder1, folder2, elnombre)
        with app.open_resource(pathx) as fp:
            msg.attach(elnombre, "application/pdf", fp.read())


    mail.send(msg)



    return 'ok'


@app.route('/<path:filename>', methods=['GET', 'POST'])
def index(filename):
    filename = filename or 'index.html'
    if request.method == 'GET':
        return send_from_directory('.', filename)


@app.route('/project')
def project():
    return render_template('editable/project.html',topmenu=topmenutemplate())


@app.route('/config_topmenu')
def edittop():
    return render_template('config/configeditor.html',topmenu=topmenutemplate())



@app.route('/kill', methods=['POST', 'GET'])
def kill():


    res = dbexecutor.executor('call Kill_Process();')

    return 'ok'

@app.route('/loadtopmenu')
def loadtopmenu():




    with open(os.path.join(app.root_path,'templates' , 'config', 'topmenu.html'), encoding='utf-8-sig') as f:
        #return APP_TEMPLATE
        datos=f.read()
        return datos


@app.route('/fileloader', methods=['POST', 'GET'])
def fileloader():

    ruta=request.args.get('archivo')

    with open(   os.path.join(app.root_path, 'templates', ruta)     , encoding='utf-8-sig') as f:
        #return APP_TEMPLATE
        datos=f.read()
        return datos

@app.route('/notificarcambioestatus', methods=['POST', 'GET'])
def notificarcambioestatus():
    idobjeto=request.args.get('idobjeto')
    tabla=request.args.get('tabla')
    idestatus=request.args.get('idestatus')

    destinatarios=[]
    destinatarios=damedestinatariosdecambioestatus(tabla,idestatus)

    valorestatus='xxxxx'

    mensaje="Notificación Cambio Estatus a " + valorestatus + " en " + tabla + " registro con id " + idobjeto + '\r\n' #+ volcadovista(tabla,idobjeto)

    for destinatario in destinatarios:
        mensaje="Notificación Cambio Estatus en " + tabla + " con id " + idobjeto
        #enviando por email el alta
        msg = Message('Notificación Cambio Estatus en ' + tabla + ' con id:' + idobjeto,
                  recipients=[destinatario])
        msg.recipients = [destinatario]
        msg.body = mensaje
        msg.html = mensaje
        mail.send(msg)






    return 'ok'

@app.route('/notificaralta', methods=['POST', 'GET'])
def notficaralta():
    idobjeto=request.args.get('idobjeto')
    tabla=request.args.get('tabla')

    destinatarios=[]
    destinatarios=damedestinatariosdealta(tabla)

    mensaje="Notificación Nuevo registro en " + tabla + " con id " + idobjeto + '\r\n' #+ volcadovista(tabla,idobjeto)

    for destinatario in destinatarios:
        mensaje="Notificación Nuevo registro en " + tabla + " con id " + idobjeto
        #enviando por email el alta
        msg = Message('Notificación Nuevo registro en ' + tabla + ' con id:' + idobjeto,
                  recipients=[destinatario])
        msg.recipients = [destinatario]
        msg.body = mensaje
        msg.html = mensaje
        mail.send(msg)


    return 'ok'





@app.route('/actualizarstatus', methods=['POST', 'GET'])
def actualizarstatus():
    tablename=request.args.get('tablename')
    campo=request.args.get('campo')
    valor=request.args.get('valor')
    idobjeto=request.args.get('idobjeto')
    tipmov=request.args.get('tipomov')

    if tipmov=='update':
        laorden="update tblestatus_" + tablename + " set " + campo + "='" + valor + "' where idestatus_" + tablename + "='" + idobjeto + "';"

    if tipmov=='insert':
        laorden="insert into tblestatus_" + tablename + " (campodisplay) values('" + valor + "');"

    res=dbexecutor.executor(laorden)

    return 'ok'



@app.route('/actualizaralta', methods=['POST', 'GET'])
def actualizaralta():
    tablename=request.args.get('tablename')
    campo=request.args.get('campo')
    valor=request.args.get('valor')
    idobjeto=request.args.get('idobjeto')
    tipmov=request.args.get('tipomov')

    if tipmov=='update':
        laorden="update tblaltanotify_" + tablename + " set " + campo + "='" + valor + "' where idaltanotify_" + tablename + "='" + idobjeto + "';"

    if tipmov=='insert':
        laorden="insert into tblaltanotify_" + tablename + " (email) values('" + valor + "');"

    res=dbexecutor.executor(laorden)

    return 'ok'



@app.route('/actualizarestatusnotify', methods=['POST', 'GET'])
def actualizarestatusnotify():
    tablename=request.args.get('tablename')
    campo=request.args.get('campo')
    valor=request.args.get('valor')
    idobjeto=request.args.get('idobjeto')
    tipmov=request.args.get('tipomov')

    if tipmov=='update':
        laorden="update tblestatusnotify_" + tablename + " set " + campo + "='" + valor + "' where idestatusnotify_" + tablename + "='" + idobjeto + "';"

    if tipmov=='insert':
        laorden="insert into tblestatusnotify_" + tablename + " (email) values('" + valor + "');"

    res=dbexecutor.executor(laorden)

    return 'ok'

@app.route('/agregarstatus', methods=['POST', 'GET'])
def agregarstatus():
    tablename=request.args.get('tablename')
    estatus=request.args.get('estatus')

    laorden="insert into tblestatus_" + tablename + " (CAMPODISPLAY) values ('" + estatus + "');"

    print(laorden)

    res=dbexecutor.executor(laorden)

    return 'ok'



@app.route('/agregaralta', methods=['POST', 'GET'])
def agregaralta():
    tablename=request.args.get('tablename')
    estatus=request.args.get('estatus')

    laorden="insert into tblaltanotify_" + tablename + " (email) values ('" + estatus + "');"

    print(laorden)
    res=dbexecutor.executor(laorden)

    return 'ok'

@app.route('/agregarestatusnotify', methods=['POST', 'GET'])
def agregarestatusnotify():
    tablename=request.args.get('tablename')
    estatus=request.args.get('estatus')
    idestatus=request.args.get('idestatus')

    laorden="insert into tblestatusnotify_" + tablename + " (email,estatus) values ('" + estatus + "','" + idestatus + "');"

    print(laorden)
    res=dbexecutor.executor(laorden)

    return 'ok'



@app.route('/eliminaralta', methods=['POST', 'GET'])
def eliminaralta():
    tablename=request.args.get('tablename')
    idestatus=request.args.get('idestatus')

    laorden="delete from tblaltanotify_" + tablename + " where idaltanotify_" + tablename + "='" + idestatus + "';"

    res=dbexecutor.executor(laorden)

    return 'ok'


@app.route('/eliminarestatusnotify', methods=['POST', 'GET'])
def eliminarestatusnotify():
    tablename=request.args.get('tablename')
    idestatus=request.args.get('idestatus')

    laorden="delete from tblestatusnotify_" + tablename + " where idestatusnotify_" + tablename + "='" + idestatus + "';"

    res=dbexecutor.executor(laorden)

    return 'ok'




@app.route('/eliminarstatus', methods=['POST', 'GET'])
def eliminarstatus():
    tablename=request.args.get('tablename')
    idestatus=request.args.get('idestatus')


    laorden="delete from tblestatus_" + tablename + " where idestatus_" + tablename + "='" + idestatus + "';"

    print(laorden)

    res=dbexecutor.executor(laorden)


    return 'ok'
@app.route('/actualizarconfigcampos', methods=['POST', 'GET'])
def actualizarconfigcampos():
    tablename=request.args.get('tablename')
    campo=request.args.get('campo')
    valor=request.args.get('valor')
    idobjeto=request.args.get('idobjeto')

    laorden="update tblconfigcampos_" + tablename + " set " + campo + "='" + valor + "' where idconfigcampos_" + tablename + "='" + idobjeto + "';"

    print(laorden)

    res=dbexecutor.executor(laorden)


    return 'ok'




@app.route('/cambiaprop', methods=['GET', 'POST'])
def cambiaprop():
    nuevoprop = request.args.get('nuevoprop')
    campokey = request.args.get('campokey')
    tablename = request.args.get('tablename')
    idobjeto = request.args.get('idobjeto')
    token = request.args.get('token')


    dbexecutor.executor('update tblprospecto set propietario=' + nuevoprop + ' where  ' + campokey + '=' + idobjeto  )

    #res2 = dbexecutor.executor("call " + tablename + "_PROC_AFTER_UPDATE(" + idobjeto + ")")
    #res3 = dbexecutor.executor("call " + tablename + "_PROC_CAMPODISPLAY(" + idobjeto + ")")

    return 'ok'



@app.route('/obtenervalor', methods=['POST', 'GET'])
def obtenervalor():
    sentencia=request.args.get('sentencia')
    print(sentencia)


    return str(resolversentencia(sentencia))

@app.route('/configcampos', methods=['POST', 'GET'])
def configcampos():
    tablename=request.args.get('tablename')
    return render_template('templateconfigcampos.html', tablename=tablename,campokey= "IDCONFIGCAMPOS_" + tablename)


def dimecampovalidacion(tablename):
    sentencia="select campo_acceso from tblperspectivaexterna where tabla_fuente='" + tablename + "'"
    return resolversentencia(sentencia)


@app.route('/loginperspectiva', methods=['POST', 'GET'])
def loginperspectiva():
    titulo = request.args.get('titulo')
    tablename=request.args.get('perspectiva')
    return render_template('loginperspectiva.html', titulo=titulo,tablename=tablename, campo_validacion=dimecampovalidacion(tablename))




@app.route('/perspectivavinculante', methods=['POST', 'GET'])
def perspectivavinculante():
    tablename=request.args.get('tabla')
    token = request.args.get('token')
    campo_acceso = request.args.get('campo_acceso')
    return render_template('loginperspectiva.html', tablename=tablename)



@app.route('/configstatus', methods=['POST', 'GET'])
def configstatus():
    tablename=request.args.get('tablename')
    return render_template('templateconfigstatus.html', tablename=tablename,campokey= "IDESTATUS_" + tablename)

@app.route('/configalta', methods=['POST', 'GET'])
def configalta():
    tablename=request.args.get('tablename')
    return render_template('templateconfigaltanotify.html', tablename=tablename,campokey= "IDALTANOTIFY_" + tablename)


@app.route('/configestatusnotify', methods=['POST', 'GET'])
def configestatusnotify():
    tablename=request.args.get('tablename')
    return render_template('templateconfigestatusnotify.html', tablename=tablename,campokey= "IDESTATUSNOTIFY_" + tablename)






@app.route('/fileloader3', methods=['POST', 'GET'])
def fileloader3():

    ruta=request.args.get('archivo')

    with open(   os.path.join(app.root_path, 'code', ruta)     , encoding='utf-8-sig') as f:
        #return APP_TEMPLATE
        datos=f.read()
        return datos



#EJEMPLO PARA FORMS
@app.route('/savetopmenu', methods=['POST', 'GET'])
def savetopmenu():

    if request.method == "POST":
        req = request.form
        datos = request.form["editor"]

        f = open(  os.path.join(app.root_path,'templates' , 'config', 'topmenu.html')  , "w",encoding='utf8')
        f.write(datos.replace("\r\n\r\n\r\n","").replace("\r\n\r\n","\r\n"))
        f.close()

    return "<!DOCTYPE html><html><head><!-- HTML meta refresh URL redirection -->   <meta http-equiv='refresh'   content='0; url=config_topmenu'></head><body></body></html>"


@app.route('/enviaradjunto', methods=['POST', 'GET'])
def enviaradjunto():

    para = request.args.get('para')
    asunto = request.args.get('asunto')
    mensaje = request.args.get('mensaje')
    nombre = request.args.get('nombre')
    reporte = unquote(request.args.get('reporte'))


    return '../../mailclient?para=' + para + '&asunto=' + asunto + '&cuerpo=' + mensaje + '&folder1=static&folder2=temp&elnombre=' + nombre + '.html' + '.pdf'

@app.route('/enviarreporte', methods=['POST', 'GET'])
def enviarreporte():

    para = request.args.get('para')
    asunto = request.args.get('asunto')
    mensaje = request.args.get('mensaje')
    formato = request.args.get('formato')
    nombre = request.args.get('nombre')
    reporte = unquote(request.args.get('reporte'))

    pathx=os.path.join(app.root_path, 'static', 'temp', nombre + '.html')
    f = open(pathx, "w")
    f.write(reporte)
    f.close()

    resultFile = open(pathx + '.pdf', "w+b")
    pisaStatus = pisa.CreatePDF(reporte,dest=resultFile)
    resultFile.close()


    return '../../mailclient?para=' + para + '&asunto=' + asunto + '&cuerpo=' + mensaje + '&folder1=static&folder2=temp&elnombre=' + nombre + '.html' + '.pdf'







@app.route('/savereport', methods=['POST', 'GET'])
def savereport():

    datos = request.args.get('datos')
    archivo = request.args.get('archivo')

    pathx=os.path.join(app.root_path, 'static', 'reports',archivo)
    f = open(pathx, "w",encoding='utf8')
    f.write(datos)
    f.close()

    return 'Saved'\


@app.route('/saveblock', methods=['POST', 'GET'])
def saveblock():

    archivo = request.args.get('archivo')
    datos = request.args.get('datos')

    print(datos)
    pathx=os.path.join(app.root_path, 'templates','tables', archivo)
    f = open(pathx, "w",encoding='utf8')
    f.write(datos)
    f.close()

    return 'Archivo Guardado' #"<!DOCTYPE html><html><head><!-- HTML meta refresh URL redirection -->   <meta http-equiv='refresh'   content='0; url=static/ace/ace.html?archivo=" + archivo + "&tipo=html'></head><body></body></html>"


@app.route('/saveblockscript', methods=['POST', 'GET'])
def saveblockscript():

    archivo = request.args.get('archivo')
    datos = request.args.get('datos')

    print(datos)
    pathx=os.path.join(app.root_path, 'static/blocky/scripts',archivo)
    f = open(pathx, "w",encoding='utf8')
    f.write(datos)
    f.close()

    return 'Archivo Guardado' #"<!DOCTYPE html><html><head><!-- HTML meta refresh URL redirection -->   <meta http-equiv='refresh'   content='0; url=static/ace/ace.html?archivo=" + archivo + "&tipo=html'></head><body></body></html>"


@app.route('/injectar', methods=['POST', 'GET'])
def injectar():

    archivo = request.args.get('archivo')
    datos = request.args.get('datos')

    print(datos)
    pathx=os.path.join(app.root_path, 'templates','tables', archivo)
    f = open(pathx, "w",encoding='utf8')
    f.write(datos)
    f.close()

    return 'Archivo Injectado'


@app.route('/injectarscript', methods=['POST', 'GET'])
def injectarscript():

    archivo = request.args.get('archivo')
    datos = request.args.get('datos')

    print(datos)
    pathx=os.path.join(app.root_path, 'static/blocky/scripts', archivo)
    f = open(pathx, "w",encoding='utf8')
    f.write(datos)
    f.close()

    return 'Archivo JS Generado'

@app.route('/getblock', methods=['POST', 'GET'])
def getblock():

    archivo = request.args.get('archivo')

    pathx=os.path.join(app.root_path, 'templates', 'tables',archivo)
    f = open(pathx, "r",encoding='utf8')
    datos=f.read()
    f.close()

    return str(datos)


@app.route('/getblockscript', methods=['POST', 'GET'])
def getblockscript():

    archivo = request.args.get('archivo')

    pathx=os.path.join(app.root_path, 'static/blocky/scripts',archivo)
    f = open(pathx, "r",encoding='utf8')
    datos=f.read()
    f.close()

    return str(datos)


@app.route('/readjs', methods=['POST', 'GET'])
def readjs():

    archivo = request.args.get('archivo')

    pathx=os.path.join(app.root_path, 'static/blocky/scripts',archivo)
    f = open(pathx, "r",encoding='utf8')
    datos=f.read()
    f.close()

    return str(datos)


@app.route('/savedashboard', methods=['POST', 'GET'])
def savedashboard():

    datos = request.args.get('datos')
    archivo = request.args.get('dashboard')

    pathx=os.path.join(app.root_path, 'static', 'dashboards',archivo)
    f = open(pathx, "w",encoding='utf8')
    f.write(datos)
    f.close()

    return 'Saved' #"<!DOCTYPE html><html><head><!-- HTML meta refresh URL redirection -->   <meta http-equiv='refresh'   content='0; url=static/ace/ace.html?archivo=" + archivo + "&tipo=html'></head><body></body></html>"


@app.route('/savetopmenu2', methods=['POST', 'GET'])
def savetopmenu2():

    datos = request.args.get('datos')
    archivo = request.args.get('archivo')

    laruta=os.path.join(app.root_path,'templates' , archivo)
    print (laruta)
    f = open(  laruta , "w",encoding='utf8')
    f.write(datos)
    f.close()

    return 'Saved' #"<!DOCTYPE html><html><head><!-- HTML meta refresh URL redirection -->   <meta http-equiv='refresh'   content='0; url=static/ace/ace.html?archivo=" + archivo + "&tipo=html'></head><body></body></html>"



@app.route('/savetopmenu3', methods=['POST', 'GET'])
def savetopmenu3():
    datos = request.args.get('datos')
    archivo = request.args.get('archivo')
    f = open(os.path.join(APP_CODE,archivo), "w",encoding='utf8')
    f.write(datos)
    f.close()
    return 'Saved' #"<!DOCTYPE html><html><head><!-- HTML meta refresh URL redirection -->   <meta http-equiv='refresh'   content='0; url=static/ace/ace.html?archivo=" + archivo + "&tipo=html'></head><body></body></html>"



@app.route('/analysis')
def analysis():
    return render_template('config/analysis.html',topmenu=topmenutemplate())




@app.route('/entities')
def entities():
    return render_template('config/entities.html',topmenu=topmenutemplate())


@app.route('/adminzone')
def adminzone():
    return render_template('config/adminzone.html',topmenu=topmenutemplate())


@app.route('/operations')
def operations():
    return render_template('config/operations.html',topmenu=topmenutemplate())


@app.route('/help')
def help():
    return render_template('config/help.html',topmenu=topmenutemplate())


@app.route('/logout')
def logout():
    return render_template('config/logout.html')


@app.route('/reports')
def reportes():
    return render_template('config/reports.html',topmenu=topmenutemplate())

@app.route('/noautorizado', methods=['POST', 'GET'])
def noautorizado():
    tablename = request.args.get('tablename')
    return render_template('config/accesonoautorizado.html',tablename=tablename)







@app.route('/insertartoken', methods=['POST', 'GET'])
def insertartoken():

    token = request.args.get('token')
    username = request.args.get('username')
    esadmin = request.args.get('esadmin')
    emailrespuesta = request.args.get('emailrespuesta')
    idusername = request.args.get('idusername')
    tipouser = request.args.get('tipouser')

    return render_template ('insertartoken.html',tipouser=tipouser,idusername=idusername,emailrespuesta=emailrespuesta,token= token,username=username,esadmin=esadmin )


@app.route('/incrustar', methods=['POST', 'GET'])
def incrustar():

    datos=''

    pagina = request.args.get('pagina')
    campo = request.args.get('campo')
    posicion = request.args.get('posicion')
    idsyscomponent = request.args.get('idsyscomponent')

    pathx = os.path.join(app.root_path, 'static', 'uploads/tblsyscomponent', idsyscomponent , 'componente.txt')
    print(pathx)

    with open(pathx, 'r', encoding='utf-8-sig') as f:
        datos=f.read();
        nombre=str(randint(1,158947))
        datos=datos.replace('_ALEA_',nombre)
        datos=datos.replace('_POS_',posicion)



    pathx = os.path.join(app.root_path, 'templates/tables', pagina + '.html')
    with open(pathx, 'a', encoding='utf-8-sig') as f:
        f.write(datos);




    pathx = os.path.join(app.root_path, 'templates', 'tables', pagina)
    #print(pathx)

    with open(pathx, 'w',encoding='utf-8-sig') as f:
       f.write(datos)

    return 'ok'




@app.route('/validartoken', methods=['POST', 'GET'])
def validartoken():
    token = request.args.get('token')

    dbexecutor.executor("call validartokens();")

    return valtoken(token)



@app.route('/soloadmin', methods=['POST', 'GET'])
def soloadmin():

    username = request.args.get('username')
    return render_template ('soloadmin.html',username=username )

@app.route('/autorizado', methods=['POST', 'GET'])
def autorizado():

    token = request.args.get('token')
    username = request.args.get('username')
    esadmin = request.args.get('esadmin')
    emailrespuesta = request.args.get('emailrespuesta')
    idusername = request.args.get('idusername')
    tipouser = dimetipo(username)

    return redirect ('/insertartoken?tipouser=' + str(tipouser) + '&idusername=' + idusername + '&emailrespuesta=' + emailrespuesta + '&esadmin=' + esadmin + '&token=' + token + '&username=' + username )



@app.route('/login')
def paglogin():
    return render_template('config/login.html')

@app.route('/checklogin', methods=['POST', 'GET'])
def checklogin():

    username=''
    password=''
    esadmin='0'

    if request.method == "POST":
        req = request.form
        username = request.form["username"]
        password = request.form["password"]

        resultado=esvalido(username,password)

        print (resultado + '***************************************************************************')

        if resultado!='No autorizado':
            #enviando token a cliente
            idusername=quieroelid(username,password)
            esadmin=checaradmin(username,password)
            emailrespuesta=checaremail(username,password)

            productividad("INICIO SESION",idusername,username,'LOGIN')

            return redirect('autorizado?emailrespuesta=' + emailrespuesta + '&esadmin=' + esadmin + '&username=' + username + '&idusername=' + str(idusername) + '&token=' + resultado )
        else:
            return redirect('login')


def elementosvinculados(tablename):
    #"/static/stimulsoft/reportdesigner.html?reporte=' + cadafile + '"

    elresultado = '<table>'
    elresultado += '<tr><td>REPORTE</td></tr>'

    elresultado +=generarelementos(tablename)


    elresultado=elresultado + '</table>'




    return elresultado


def esvalidopers(tablename,codigo, campo_validacion):



    resultado="Incorrecto"
    lasentencia="select count(*) as total from  " + tablename + "  where " + campo_validacion + "='" +  codigo + "'   ;"

    print(lasentencia)

    validacion=resolversentencia(lasentencia)

    print('resultado --> >>>>>>>>>>>>>>>>>>>> ' + str(validacion))

    if  validacion==1 or validacion=='1'   :
        resultado="Correcto"
        return render_template("templateperspectiva.html", elementosvinculados = str(elementosvinculados(tablename)).replace('_XXX_',codigo) )

    if validacion==0 or validacion=='0' or validacion=='None'   :
        resultado="Incorrecto"
        return render_template("config/accesonoautorizado.html")





@app.route('/checkloginpers', methods=['POST', 'GET'])
def checkloginpers():

    tablename = request.args.get('tabla')
    campo_validacion = request.args.get('campo')


    print(tablename + campo_validacion + "************************************")

    if request.method == "POST":
        req = request.form
        codigo = request.form["codigo"]

        resultado=esvalidopers(tablename,codigo,campo_validacion)

        return resultado




@app.route('/jqwidget')
def jqwid():
    return render_template('samples/jqwidget.html',topmenu=topmenutemplate())









@app.route('/ws')
def muestralibros():
    return jsonify(libreria.books)



@app.route('/pagina2')
def home2():
    now = datetime.now()
    formatted_now = now.strftime("%A, %d %B, %Y at %X")
    cadena='<b> Hello, Flask! </b> on '

    return render_template("index.html",  content = cadena + formatted_now)

#crea nuevo token
@app.route('/creatoken', methods=['GET','POST'])
def creatoken():
    username = request.args.get('username')
    password= request.args.get('password')
    return username + " " + password






@app.route('/damereportes', methods=['GET','POST'])
def damereportes():

    pathx=os.path.join(app.root_path, 'static', 'reports')
    onlyfiles = [f for f in listdir(pathx) if isfile(join(pathx, f))]

    losfiles='<table><tr><td>REPORTE</td><td>Diseñar</td><td>Visualizar</td><td>Eliminar</td></tr>'
    for cadafile in onlyfiles:
        losfiles +='<tr><td>' + cadafile + '</td><td> <a href="/static/stimulsoft/reportdesigner.html?reporte=' + cadafile + '" >Diseñar</a> </td><td>  <a href="/static/stimulsoft/reportviewer.html?reporte=' + cadafile + '" >Visualizar</a> </td><td> <a href="/static/stimulsoft/reportdesigner.html?reporte=' + cadafile + '" >Eliminar</a> </td></tr>'

    elresultado=losfiles + '</table>'
    return render_template("templatereportes.html",filas=elresultado)

@app.route('/dameblocks', methods=['GET','POST'])
def dameblocks():

    pathx=os.path.join(app.root_path, 'static', 'blocky/scripts')
    onlyfiles = [f for f in listdir(pathx) if isfile(join(pathx, f))]

    losfiles=''
    for cadafile in onlyfiles:
        losfiles += '   *' + cadafile

    elresultado=losfiles
    return  elresultado


@app.route('/damedashboards', methods=['GET','POST'])
def damedashboards():

    pathx=os.path.join(app.root_path, 'static', 'dashboards')
    onlyfiles = [f for f in listdir(pathx) if isfile(join(pathx, f))]

    losfiles='<table><tr><td>DASHBOARD</td><td>Diseñar</td><td>Visualizar</td><td>Eliminar</td></tr>'
    for cadafile in onlyfiles:
        losfiles +='<tr><td>' + cadafile + '</td><td> <a href="/static/stimulsoft/dashboarddesigner.html?dashboard=' + cadafile + '" >Diseñar</a> </td><td>  <a href="/static/stimulsoft/dashboardviewer.html?dashboard=' + cadafile + '" >Visualizar</a> </td><td> <a href="/static/stimulsoft/reportdesigner.html?reporte=' + cadafile + '" >Eliminar</a> </td></tr>'

    elresultado=losfiles + '</table>'
    return render_template("templatedashboards.html",filas=elresultado)


#ejecuta codigo SQL
@app.route('/sqlexec2', methods=['GET','POST'])
def sqlexec2():
    sentencia = request.args.get('sentencia')
    tablename= request.args.get('tablename')
    idusername= request.args.get('idusername')

    sentencia=sentencia.replace("''","Null")

    resultado=dbexecutor.executor2(sentencia,tablename)

    nuevoobjeto=dameultimoid(tablename)
    res2=dbexecutor.executor("call " + tablename + "_PROC_AFTER_INSERT(" + nuevoobjeto + ");")
    res3=dbexecutor.executor("call " + tablename + "_PROC_CAMPODISPLAY(" + nuevoobjeto + ")")


    #asignando el propietario
    res4=dbexecutor.executor("update  " + tablename + " set propietario='" + str(idusername)  + "' where  ID" + tablename.replace("tbl","") + "=" + nuevoobjeto )



    return str(nuevoobjeto)

#ejecuta codigo SQL
@app.route('/sqlexec', methods=['GET','POST'])
def sqlexec():
    idobjeto= request.args.get('idobjeto')
    tablename= request.args.get('tablename')
    sentencia = request.args.get('sentencia')
    token= request.args.get('token')
    print(sentencia)

    sentencia=sentencia.replace("=''","=Null")

    sentencia=sentencia.replace("='undefined'","=Null")
    dbexecutor.executor(sentencia)

    res2=dbexecutor.executor("call " + tablename + "_PROC_AFTER_UPDATE(" + idobjeto + ")")
    res3=dbexecutor.executor("call " + tablename + "_PROC_CAMPODISPLAY(" + idobjeto + ")")

    return 'ok'


@app.route('/testjson', methods=['GET','POST'])
def testjson():
    tablename = request.args.get('tablename')
    damecampos=''
    for distro in getfields2(tablename):
        damecampos = damecampos + '0,DATOS,0,' + distro + ',' + distro + '\r'

    return damecampos

@app.route('/resetpopulationquery', methods=['GET', 'POST'])
def resetpopulationquery():
    table = request.args.get('table')
    pathx = os.path.join(app.root_path, 'templates/tables', table)
    os.remove(pathx)

    return 'ok'






@app.route('/deletefile', methods=['GET','POST'])
def deletefile():
    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')
    filename= request.args.get('filename')

    pathx=os.path.join(app.root_path, 'static', 'uploads',tabla,idobjeto,filename)
    os.remove(pathx)


    return 'ok'
#edita un registro
@app.route('/refrescador', methods=['GET','POST'])
def refrescador():
    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')

    res=dbexecutor.executor("call " + tabla + "_PROC_REFRESCADOR(" + idobjeto + ");")


    return 'ok'








@app.route('/damefilesslider', methods=['GET','POST'])
def damefilesslider():
    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')


    pathx=os.path.join(app.root_path, 'static', 'uploads',tabla,idobjeto)
    onlyfiles = [f for f in listdir(pathx) if isfile(join(pathx, f))]


    contador=0
    losfiles=''
    for cadafile in onlyfiles:
        if contador==0:
            if ".png" in cadafile.lower()  or ".jpg" in cadafile.lower()  or ".bmp" in cadafile.lower()  or ".gif" in cadafile.lower() or ".jpeg" in cadafile.lower()    :
                losfiles +='<img style="width:300px;position:relative;left:-60px"    src="/static/uploads/' + tabla + '/' + idobjeto + '/' + cadafile + '" />\r\n'
                contador +=1

                dbexecutor.executor("update " + tabla + " set imagenadjunta='" + cadafile + "' where ID" +  (tabla.replace("tbl","")).upper()  + "='" + idobjeto + "'")


    return losfiles + ''


@app.route('/savenota', methods=['GET', 'POST'])
def savenota():

    ruta = request.args.get('ruta')
    datos = request.args.get('datos')

    pathfinal=ruta.replace("*","/")


    pathx = os.path.join(app.root_path, 'static', 'notas', pathfinal,'notas.html')
    print(pathx)
    with open(pathx, 'w',encoding='utf-8-sig') as f:
       f.write(datos)


    return 'ok'

@app.route('/guardardiagrama', methods=['GET', 'POST'])
def guardardiagrama():

    eldiagrama = request.args.get('diagrama')
    datos = request.args.get('datos')


    pathx = os.path.join(app.root_path, 'static', 'diagramas', 'diagramas', eldiagrama)
    with open(pathx, 'w',encoding='utf-8-sig') as f:
       f.write(datos)


    return 'ok'


@app.route('/deleteperspectiva', methods=['GET', 'POST'])
def deleteperspectiva():

    filename = request.args.get('filename')
    query = request.args.get('query')




    pathx = os.path.join(app.root_path, 'static', 'pivot', 'perspectiva',query ,filename)

    os.remove(pathx)



    return 'ok'

@app.route('/deletediagram', methods=['GET', 'POST'])
def deletediagram():

    filename = request.args.get('filename')


    pathx = os.path.join(app.root_path, 'static', 'diagramas', 'diagramas', filename)

    os.remove(pathx)



    return 'ok'

@app.route('/setperspectiva', methods=['GET', 'POST'])
def setperspectiva():

    query = request.args.get('query')
    perspectiva = request.args.get('perspectiva')
    datos = str(request.data)

    #creando la carpeta del query
    pathx = os.path.join(app.root_path, 'static', 'pivot','perspectiva' ,query)
    if not os.path.isdir(pathx):
        os.mkdir(pathx)

    #guardando/actualizando la perspectiva
    pathx = os.path.join(app.root_path, 'static', 'pivot','perspectiva' ,query,perspectiva)

    if os.path.isfile(pathx):
        os.remove(pathx)

    if not os.path.isfile(pathx):
        f = open(  pathx  , "w",encoding='utf8')
        f.write(datos)
        f.close()


    return 'Perspectiva Guardada/Actualizada'





@app.route('/getperspectiva', methods=['GET', 'POST'])
def getperspectiva():

    query = request.args.get('query')
    perspectiva = request.args.get('perspectiva')

    pathx = os.path.join(app.root_path, 'static', 'pivot','perspectiva' ,query,perspectiva)

    with open(pathx, encoding='utf-8-sig') as f:
        datos=f.read()

    print(datos);

    return jsonify(datos)




@app.route('/cargarnota', methods=['GET', 'POST'])
def cargarnota():

    ruta = request.args.get('ruta')
    lanota=ruta.replace("*","/") + "/notas.html"


    pathx = os.path.join(app.root_path, 'static', 'notas', lanota)

    print(pathx + 'ddddddddddddddddddd')

    with open(pathx, encoding='utf-8-sig') as f:
        datos=f.read()


    return datos

@app.route('/cargardiagrama', methods=['GET', 'POST'])
def cargardiagrama():

    eldiagrama = request.args.get('diagrama')
    datos=''

    pathx = os.path.join(app.root_path, 'static', 'diagramas', 'diagramas', eldiagrama)
    with open(pathx, encoding='utf-8-sig') as f:
        datos=f.read()


    return datos

@app.route('/dameperspectivas', methods=['GET', 'POST'])
def dameperspectivas():


    query = request.args.get('query')


    pathx = os.path.join(app.root_path, 'static', 'pivot', 'perspectiva', query)
    onlyfiles = [f for f in listdir(pathx) if isfile(join(pathx, f))]

    losfiles = '<table style="border: 1px solid black;width:550px"   ><tr><td style="background-color:lightblue;margin-right:20px;border: 1px solid black;" >PERSPECTIVA</td> <td  style="background-color:lightblue;margin-right:20px;border: 1px solid black;" >MOSTRAR</td> <td  style="background-color:lightblue;margin-right:20px;border: 1px solid black;">ELIMINAR</td></tr>'
    for cadafile in onlyfiles:

        losfiles += '<tr><td style="margin-right:20px;border: 1px solid black;" >' + cadafile + '</td><td style="margin-right:20px;border: 1px solid black;"  > <a  target="_self" href="/static/pivot/index.html?query=' + query + '&perspectiva=' + cadafile + '" >' + 'Ver Perspectiva' + '</a> </td><td  style="margin-right:20px;border: 1px solid black;"   > <a href=''javascript:deldiagram("deleteperspectiva?query=' + query + '&filename=' + cadafile.replace(' ','%20') + '") >Eliminar Perspectiva</a> </td></tr>'

    return losfiles + '</table>'

@app.route('/damediagramas', methods=['GET', 'POST'])
def damediagramas():

    pathx = os.path.join(app.root_path, 'static', 'diagramas', 'diagramas')
    onlyfiles = [f for f in listdir(pathx) if isfile(join(pathx, f))]

    losfiles = '<table style="border: 1px solid black;width:550px"   ><tr><td style="background-color:lightblue;margin-right:20px;border: 1px solid black;" >DIAGRAMA</td> <td  style="background-color:lightblue;margin-right:20px;border: 1px solid black;" >EDITAR</td> <td  style="background-color:lightblue;margin-right:20px;border: 1px solid black;">ELIMINAR</td></tr>'
    for cadafile in onlyfiles:

        losfiles += '<tr><td style="margin-right:20px;border: 1px solid black;" >' + cadafile + '</td><td style="margin-right:20px;border: 1px solid black;"  > <a  target="_self" href="/static/diagramas/index.html?diagrama=' + cadafile + '" >Editar Diagrama</a> </td><td  style="margin-right:20px;border: 1px solid black;"   > <a href=''javascript:deldiagram("deletediagram?filename=' + cadafile + '"'') >Eliminar Digrama</a> </td></tr>'

    return losfiles + '</table>'


@app.route('/damefiles', methods=['GET','POST'])
def damefiles():
    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')

    #pregenerando las carpetas de notas y adjuntos
    pathx = os.path.join(app.root_path, 'static', 'notas', tabla)
    if not os.path.isdir(pathx):
        os.mkdir (pathx)

    pathx = os.path.join(app.root_path, 'static', 'notas', tabla, idobjeto)
    if not os.path.isdir(pathx):
        os.mkdir (pathx)

    pathx = os.path.join(app.root_path, 'static', 'uploads', tabla)
    if not os.path.isdir(pathx):
        os.mkdir (pathx)

    pathx = os.path.join(app.root_path, 'static', 'uploads', tabla, idobjeto)
    if not os.path.isdir(pathx):
        os.mkdir (pathx)

    pathx = os.path.join(app.root_path, 'static', 'notas', tabla, idobjeto,'notas.html')
    if not os.path.isfile(pathx):
        f = open(  pathx  , "w",encoding='utf8')
        f.write('')
        f.close()




    pathx=os.path.join(app.root_path, 'static', 'uploads',tabla,idobjeto)
    onlyfiles = [f for f in listdir(pathx) if isfile(join(pathx, f))]

    losfiles='<table style="border: 1px solid black;width:550px"   ><tr><td  style="background-color:lightblue;margin-right:20px;border: 1px solid black;">ELIMINAR</td><td style="background-color:lightblue;margin-right:20px;border: 1px solid black;"  >IMAGEN</td><td style="background-color:lightblue;margin-right:20px;border: 1px solid black;" >ARCHIVO</td> <td  style="background-color:lightblue;margin-right:20px;border: 1px solid black;" >DESCARGAR</td> <td style="background-color:lightblue;margin-right:20px;border: 1px solid black;" >ELIMINAR</td></tr>'
    for cadafile in onlyfiles:

        if ".png" in cadafile.lower()  or ".jpg" in cadafile.lower()  or ".bmp" in cadafile.lower()  or ".gif" in cadafile.lower() or ".jpeg" in cadafile.lower()    :
            losfiles +='<tr><td  style="margin-right:20px;border: 1px solid black;"   > <a href=''javascript:delfile("deletefile?mobile=1&tabla=' + tabla + '&idobjeto=' + idobjeto + '&filename=' + cadafile + '"'') >Eliminar</a> </td><td style="margin-right:20px;border: 1px solid black;" >  <img alt="No es imagen" src="/static/uploads/' + tabla + '/' + idobjeto + '/' + cadafile + '"   height="80px" /> </td><td style="margin-right:20px;border: 1px solid black;" >' + cadafile + '</td><td  style="margin-right:20px;border: 1px solid black;" > <a target="_self" href="/static/uploads/' + tabla + '/' + idobjeto + '/' + cadafile + '" >Archivo</a> </td><td  style="margin-right:20px;border: 1px solid black;"   > <a target="_self" href="../../mailclient?asunto=Envio de archivo adjunto&cuerpo=Estimado client@, le envio archivo adjunto...&folder1=static&folder2=uploads/' + tabla + '/' + idobjeto + '&elnombre=' + cadafile + '&para="    >Enviar por Email</a>  </td></tr>'
        else:
            losfiles +='<tr><td style="margin-right:20px;border: 1px solid black;" >' + cadafile + '</td><td style="margin-right:20px;border: 1px solid black;"  > <a  target="_self" href="/static/uploads/' + tabla + '/' + idobjeto + '/' + cadafile + '" >Archivo</a> </td><td style="margin-right:20px;border: 1px solid black;" >   </td><td  style="margin-right:20px;border: 1px solid black;"   > <a href=''javascript:delfile("deletefile?tabla=' + tabla + '&idobjeto=' + idobjeto + '&filename=' + cadafile + '"'') >Eliminar</a> </td><td   > <a target="_self" href="../../mailclient?asunto=Envio de archivo adjunto&cuerpo=Estimado client@, le envio archivo adjunto...&folder1=static&folder2=uploads/' + tabla + '/' + idobjeto + '&elnombre=' + cadafile + '&para="     >Enviar por Email</a>  </td></tr>'


    return losfiles + '</table><br><h5>Notas</h5> <iframe id="lasnotas" style="width: 95%;height: 700px;overflow: hidden"  src="/static/editor/index.html?ruta=' + tabla + '*' + idobjeto + '"  /> </iframe> <br> <a target="_blank" href="/static/notas/' + tabla + '/' + idobjeto + '/notas.html' + '"' + ' >URL Directa</a>'

@app.route('/formrender', methods=['GET','POST'])
def formrender():
    formx = request.args.get('form')
    idobjeto= request.args.get('idobjeto')
    token= request.args.get('token')
    return render_template(formx + '.html')



@app.route('/voiceaction', methods=['GET','POST'])
def voiceaction():
    voice = request.args.get('voice')

    partes=[]
    partes=voice.split(" ")

    if partes[0].lower()=='buscar':
        latabla="tbl" + partes[1].lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
        salida="/rendergrid?grid=tables/" + latabla  + "&buscar=" + partes[2]
        print ("salida --> " + salida)
        return salida


    return url_for(voice)


@app.route('/restauracionmasiva', methods=['GET','POST'])
def restauracionmasiva():

    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')

    res=dbexecutor.executor("update " + tabla + " set eliminado='false' where ID" + tabla.replace('tbl','') + "='" + idobjeto + "';")



    return 'ok'

@app.route('/eliminacionmasiva', methods=['GET','POST'])
def eliminacionmasiva():

    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')

    res=dbexecutor.executor("update " + tabla + " set eliminado='true' where ID" + tabla.replace('tbl','') + "='" + idobjeto + "';")



    return 'ok'

@app.route('/bloqueomasivo', methods=['GET','POST'])
def bloqueomasivo():

    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')

    res=dbexecutor.executor("update " + tabla + " set bloqueado='true' where ID" + tabla.replace('tbl','') + "='" + idobjeto + "';")



    return 'ok'

@app.route('/bloqueototal', methods=['GET','POST'])
def bloqueototal():

    tabla = request.args.get('tabla')


    res=dbexecutor.executor("update " + tabla + " set bloqueado='true';")



    return 'ok'



@app.route('/desbloqueomasivo', methods=['GET','POST'])
def desbloqueomasivo():

    tabla = request.args.get('tabla')
    idobjeto= request.args.get('idobjeto')

    res=dbexecutor.executor("update " + tabla + " set bloqueado='false' where ID" + tabla.replace('tbl','') + "='" + idobjeto + "';")



    return 'ok'

@app.route('/renderconfig', methods=['GET','POST'])
def configrender():
    formx = request.args.get('argument')
    return render_template(formx + '.html')

@app.route('/rendergrid', methods=['GET','POST'])
def rendergrid():
    formx = request.args.get('grid')

    return render_template(formx + '.html')



@app.route('/actualizartabla', methods=['GET','POST'])
def actualizartabla():

    tabla = request.args.get('tablename')
    idobjeto = request.args.get('idobjeto')
    campo = request.args.get('campo')
    valor = request.args.get('valor')
    usuario = request.args.get('usuario')
    idusername = request.args.get('idusername')



    sentenciaupdate=""

    campokey="ID" + tabla.replace("tbl","")


    sentenciaupdate="update " + tabla + " set " + campo + " = '" + valor + "' where ID" + tabla.replace("tbl","") + "=" + idobjeto + " ;"

    sentenciaupdate = sentenciaupdate.replace("_aptphe_","''")
    sentenciaupdate = sentenciaupdate.replace("_and_", "&")

    sentenciaupdate = sentenciaupdate.replace("=''","=Null")

    sentenciaupdate = sentenciaupdate.replace("='null'","=Null")
    sentenciaupdate = sentenciaupdate.replace("='None'","=Null")


    print (sentenciaupdate)



    #REGISTRO DE PRODUCTIVIDAD
    productividad("ACTUALIZACION REGISTRO " + str(idobjeto),idusername,usuario,tabla)

    res=dbexecutor.executor(sentenciaupdate)
    if res==1:
        #ejecutando after update
        res2=dbexecutor.executor("call " + tabla + "_PROC_AFTER_UPDATE(" + idobjeto + ")")

        print("call " + tabla + "_PROC_CAMPODISPLAY(" + idobjeto + ")" + '**********************************************************************************************')
        res3=dbexecutor.executor("call " + tabla + "_PROC_CAMPODISPLAY(" + idobjeto + ")")


        print("update " + tabla + " set ultimousuario='" + str(usuario) + "', horamodificacion=time(now()) where "  +  campokey + "=" + str(idobjeto) + 'pppppppppppppppppppppppppppppppppppppppppppppppppppppp')

        res4=dbexecutor.executor("update " + tabla + " set ultimousuario='" + str(usuario) + "', horamodificacion=time(now()) where "  +  campokey + "=" + str(idobjeto) )

        res5=dbexecutor.executor("DELETE FROM " + tabla + "_history WHERE fechamodificacion < NOW() - INTERVAL 90 DAY;" )

    return sentenciaupdate





@app.route('/saverecord', methods=['GET','POST'])
def saverecord():
    objeto = request.args.get('objeto')
    tipomov = request.args.get('tipomov')
    tabla = request.args.get('tabla')
    idobjeto = request.args.get('idobjeto')
    campokey = request.args.get('campokey')
    usuario = request.args.get('usuario')
    idusername = request.args.get('idusername')

    print (tipomov)


    #objeto=objeto[:-1]
    #objeto=objeto[1:]

    objetos=[]
    objetos=json.loads(objeto)

    loscampos=''
    losvalores=''

    sentenciainsert=""
    sentenciaupdate=""
    cadenadeupdates=""

    for cadaobjeto in objetos:
        for key, value in cadaobjeto.items():

            if tipomov=='nuevo':
                if campokey!=key:
                    loscampos=loscampos + key + ","
                    losvalores=losvalores + "'" +  value + "',"

            if tipomov=='update':
                if campokey!=key:
                    cadenadeupdates=cadenadeupdates + key +  "='" + value + "',"



    if tipomov=='update':
        cadenadeupdates=cadenadeupdates[:-1]

        sentenciaupdate="update " + tabla + ' set ' + cadenadeupdates + ' where ' + campokey + '=' + idobjeto + ' ;'

        sentenciaupdate = sentenciaupdate.replace("_aptphe_","''")
        sentenciaupdate = sentenciaupdate.replace("_and_","&")

        sentenciaupdate = sentenciaupdate.replace("=''","=Null")

        sentenciaupdate = sentenciaupdate.replace("='null'","=Null")
        sentenciaupdate = sentenciaupdate.replace("='None'","=Null")


        print (sentenciaupdate)



        #REGISTRO DE PRODUCTIVIDAD
        productividad("ACTUALIZACION REGISTRO " + str(idobjeto),idusername,usuario,tabla)

        res=dbexecutor.executor(sentenciaupdate)
        if res==1:
            #ejecutando after update
            res2=dbexecutor.executor("call " + tabla + "_PROC_AFTER_UPDATE(" + idobjeto + ")")

            print("call " + tabla + "_PROC_CAMPODISPLAY(" + idobjeto + ")" + '**********************************************************************************************')
            res3=dbexecutor.executor("call " + tabla + "_PROC_CAMPODISPLAY(" + idobjeto + ")")


            print("update " + tabla + " set ultimousuario='" + str(usuario) + "', horamodificacion=time(now()) where "  +  campokey + "=" + str(idobjeto) + 'pppppppppppppppppppppppppppppppppppppppppppppppppppppp')

            res4=dbexecutor.executor("update " + tabla + " set ultimousuario='" + str(usuario) + "', horamodificacion=time(now()) where "  +  campokey + "=" + str(idobjeto) )

            res5=dbexecutor.executor("DELETE FROM " + tabla + "_history WHERE fechamodificacion < NOW() - INTERVAL 90 DAY;" )

            return sentenciaupdate
        else:
            return res




    if tipomov=='nuevo':
        loscampos=loscampos[:-1]
        losvalores=losvalores[:-1]

        sentenciainsert="insert into " + tabla + '(' + loscampos + ') values(' + losvalores + ');'

        sentenciainsert = sentenciainsert.replace("_aptphe_","''")
        sentenciainsert = sentenciainsert.replace("_and_", "&")


        sentenciainsert=sentenciainsert.replace("''","Null")
        sentenciainsert = sentenciainsert.replace("='null'","=Null")
        sentenciainsert = sentenciainsert.replace("='None'","=Null")

        print(sentenciainsert + '**********************************************************************************')





        res=dbexecutor.executor(sentenciainsert)
        if res==1:
            nuevoobjeto=dameultimoid(tabla)

            #REGISTRO DE PRODUCTIVIDAD
            productividad("NUEVO REGISTRO " + str(nuevoobjeto),idusername,usuario,tabla)

            res2=dbexecutor.executor("call " + tabla + "_PROC_AFTER_INSERT(" + nuevoobjeto + ");")
            res3=dbexecutor.executor("call " + tabla + "_PROC_CAMPODISPLAY(" + nuevoobjeto + ")")



            #asignando el propietario
            res4=dbexecutor.executor("update  " + tabla + " set propietario='" + str(idusername)  + "' where  ID" + tabla.replace("tbl","") + "=" + nuevoobjeto )

            res5=dbexecutor.executor("update " + tabla + " set ultimousuario='" + str(usuario) + "', horamodificacion=time(now())where  ID" + tabla.replace("tbl","") + "=" + nuevoobjeto )


            res6=dbexecutor.executor("DELETE FROM " + tabla + "_history WHERE fechamodificacion < NOW() - INTERVAL 90 DAY;" )


            return sentenciainsert
        else:
            return res








@app.route('/getidobject', methods=['GET','POST'])
def getidobject():

    dbexecutor.executor('call Kill_Process();')

    tabla = request.args.get('tabla')
    idobjeto = request.args.get('idobjeto')

    #obteniendo arreglo Json de los datos de ese registro
    return  damedatosobjeto(tabla,idobjeto)




if __name__ == '__main__':
    import os
    HOST = os.environ.get('SERVER_HOST', 'localhost')
    try:
        PORT = int(os.environ.get('SERVER_PORT', '5555'))
    except ValueError:
        PORT = 5555
    app.run(HOST, PORT)


def quierolinks():

    return 'ok'



@app.route('/builderscript', methods=['GET','POST'])
def builderscript():

    return render_template("renderscript.html", resultado=tablasparabuildscript())



@app.route('/builder', methods=['GET','POST'])
def builder():
    tablename = request.args.get('tablename')

    resultado=''

    #Building Process.
    #Step 1. Table adaptation
    resultado=builderadaptation.builderadaptation(tablename)


    #Step 2. Grid Build
    buildergrid.buildergrid(tablename)

    #Step 3. Config Builder
    #builderconfigurator.buildconfigurator(tablename)

    #Step 4. Form Build
    builderform.builderform(tablename)


    #Step 5. Form Build Google Script
    googlebuilder.builderform(tablename)

    return 'ok'



@app.route('/quierodisplay', methods=['GET','POST']) #image.save(os.path.join(app.root_path, 'static', 'uploads', image.filename))
def quierodisplay():
    tablename = request.args.get('tablename')
    idobjeto = request.args.get('idobjeto')
    return damedisplay(tablename,idobjeto)





@app.route('/upload', methods=['GET','POST']) #image.save(os.path.join(app.root_path, 'static', 'uploads', image.filename))
def upload_file():
    tabla = request.args.get('tabla')
    objeto = request.args.get('objeto')
    vistadetalle = request.args.get('vistadetalle')
    campo1 = request.args.get('campo1')
    valor1 = request.args.get('valor1')
    mobile = request.args.get('mobile')


    # Get the name of the uploaded files
    uploaded_files = request.files.getlist("file[]")
    filenames = []
    for file in uploaded_files:
        # Check if the file is one of the allowed types/extensions
        if file and allowed_file(file.filename):
            # Make the filename safe, remove unsupported chars
            filename = secure_filename(file.filename)
            # Move the file form the temporal folder to the upload
            # folder we setup

            pathx=os.path.join(app.root_path, 'static', 'uploads',tabla)
            if not path.exists(pathx):
                print(pathx)
                os.mkdir(pathx)

            pathx=os.path.join(app.root_path, 'static', 'uploads',tabla,objeto)
            if not path.exists(pathx):
                print(pathx)
                os.mkdir(pathx)


            file.save(os.path.join(app.root_path, 'static', 'uploads',tabla,objeto, filename))
            # Save the filename into a list, we'll use it later
            filenames.append(filename)
            # Redirect the user to the uploaded_file route, which
            # will basicaly show on the browser the uploaded file
    # Load an html page with a link to each uploaded file
    rutanormal='formrender?campo1=' + campo1 + '&valor1=' + valor1 + '&vistadetalle=' + vistadetalle + '&form=tables/' + tabla + '_form&idobjeto=' + objeto + '&viewmode=files'
    rutamobile='formrender?campo1=' + campo1 + '&valor1=' + valor1 + '&vistadetalle=' + vistadetalle + '&form=tables/' + tabla + '_form_mobile&idobjeto=' + objeto + '&viewmode=files'

    if mobile=="1":
        return redirect(rutamobile)
    else:
        return redirect(rutanormal)


def allowed_file(filename):
	return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



@app.route('/agenda')
def agenda():
    return render_template('agenda.html',topmenu=topmenutemplate())

@app.route('/agendasave', methods=['GET','POST'])
def agendasave():

     modo = request.args.get('modo')
     id = request.args.get('id')
     subject = request.args.get('subject')
     location = request.args.get('location')
     fromx = request.args.get('from')
     to = request.args.get('to')
     status = request.args.get('status')
     description = request.args.get('description')

     campofiltro = request.args.get('campofiltro')
     idobjetofiltro = request.args.get('idobjetofiltro')

     allDay = request.args.get('allDay')
     background = request.args.get('background')
     color = request.args.get('color')
     draggable = request.args.get('draggable')
     hidden = request.args.get('hidden')
     recurrencePattern = request.args.get('recurrencePattern')
     recurrenceException = request.args.get('recurrenceException')
     resizable = request.args.get('resizable')
     resourceId = request.args.get('resourceId')
     readOnly = request.args.get('readOnly')
     tooltip = request.args.get('tooltip')
     timeZone = request.args.get('timeZone')


     if allDay==None or allDay=='null' :
         allDay=''
     if background==None or background=='null' :
         background=''
     if color==None or color=='null':
         color=''
     if draggable==None or draggable=='null' :
         draggable=''
     if recurrencePattern==None or recurrencePattern=='null' :
         recurrencePattern=''
     if recurrenceException==None or recurrenceException=='null' :
         recurrenceException=''
     if resizable==None or resizable=='null' :
         resizable=''
     if resourceId==None or resourceId=='null' :
         resourceId=''
     if readOnly==None or readOnly=='null' :
         readOnly=''
     if tooltip==None or tooltip=='null' :
         tooltip=''
     if timeZone==None or timeZone=='null' :
         timeZone=''

     sentencia=''

     ejecutado=0

     if modo=='insert':

        if campofiltro!='null' and campofiltro!=None and campofiltro!='null':
                sentencia="insert into tblagenda(" + campofiltro + ",id,name,adress,start,end,status,about,allDay,background,color,recurrencePattern,recurrenceException,resizable,resourceId,tooltip,timeZone) values('" + idobjetofiltro + "','" + id + "','" + subject + "','" + location + "','" + fromx + "','" + to + "','" + status + "','" + description + "','" + allDay + "','" + background + "','" + color + "','" + recurrencePattern + "','" + recurrenceException + "','" + resizable + "','" + resourceId + "','" + tooltip + "','" + timeZone + "');"
                ejecutado=1

        if ejecutado==0:
                sentencia="insert into tblagenda(id,name,adress,start,end,status,about,allDay,background,color,recurrencePattern,recurrenceException,resizable,resourceId,tooltip,timeZone) values('" + id + "','" + subject + "','" + location + "','" + fromx + "','" + to + "','" + status + "','" + description + "','" + allDay + "','" + background + "','" + color + "','" + recurrencePattern + "','" + recurrenceException + "','" + resizable + "','" + resourceId + "','" + tooltip + "','" + timeZone + "');"


        sentencia=sentencia.replace("=''",'=Null')
        sentencia=sentencia.replace("=' '",'=Null')
        sentencia=sentencia.replace("='  '",'=Null')

     if modo=='delete':
         sentencia="delete from tblagenda where id='" + id + "';"
         sentencia=sentencia.replace("=''",'=Null')
         sentencia=sentencia.replace("=' '",'=Null')
         sentencia=sentencia.replace("='  '",'=Null')

     if modo=='update':
         sentencia="update tblagenda set allDay='" + allDay + "',background='" + background + "',color='" + color + "',recurrencePattern='" + recurrencePattern + "',recurrenceException='" + recurrenceException + "',resizable='" + resizable + "',resourceId='" + resourceId + "',tooltip='" + tooltip + "',timeZone='" + timeZone + "', name='" + subject + "',adress='" + location + "',start='" + fromx + "',end='" + to + "',status='" + status + "',about='" + description + "' where id='" + id + "';"
         sentencia=sentencia.replace("=''",'=Null')
         sentencia=sentencia.replace("=' '",'=Null')
         sentencia=sentencia.replace("='  '",'=Null')



     dbexecutor.executor(sentencia)
     dbexecutor.executor('call agenda_ajustar_datos();')



     sentencia="update tblagenda t1 set t1.DESDE_FECHA=concat(		mid(START, 12, 4),		'-',		CASE	WHEN mid(START, 5, 3) = 'Jan' THEN		1	WHEN mid(START, 5, 3) = 'Feb' THEN		2	WHEN mid(START, 5, 3) = 'Mar' THEN		3	WHEN mid(START, 5, 3) = 'Apr' THEN		4	WHEN mid(START, 5, 3) = 'May' THEN		5	WHEN mid(START, 5, 3) = 'Jun' THEN		6	WHEN mid(START, 5, 3) = 'Jul' THEN		7	WHEN mid(START, 5, 3) = 'Ago' THEN		8	WHEN mid(START, 5, 3) = 'Sep' THEN		9	WHEN mid(START, 5, 3) = 'Oct' THEN		10	WHEN mid(START, 5, 3) = 'Nov' THEN		11	WHEN mid(START, 5, 3) = 'Dec' THEN		12	END,	'-',	mid(START, 9, 2)), DESDE_HORA=TIME_FORMAT(mid(START, 17, 8), '%H:%i:%s %p') ,HASTA_FECHA=concat(		mid(START, 12, 4),		'-',		CASE	WHEN mid(end, 5, 3) = 'Jan' THEN		1	WHEN mid(end, 5, 3) = 'Feb' THEN		2	WHEN mid(end, 5, 3) = 'Mar' THEN		3	WHEN mid(end, 5, 3) = 'Apr' THEN		4	WHEN mid(end, 5, 3) = 'May' THEN		5	WHEN mid(end, 5, 3) = 'Jun' THEN		6	WHEN mid(end, 5, 3) = 'Jul' THEN		7	WHEN mid(end, 5, 3) = 'Ago' THEN		8	WHEN mid(end, 5, 3) = 'Sep' THEN		9	WHEN mid(end, 5, 3) = 'Oct' THEN		10	WHEN mid(end, 5, 3) = 'Nov' THEN		11	WHEN mid(end, 5, 3) = 'Dec' THEN		12	END,	'-',	mid(end, 9, 2)),HASTA_HORA=TIME_FORMAT(mid(end, 17, 8), '%H:%i:%s %p')   where id='" + id + "'; "
     dbexecutor.executor(sentencia)

     print(sentencia)
     return sentencia

@app.route('/generartemplate', methods=['GET','POST'])
def generartemplate():

    tablename = request.args.get('tablename')
    idobjeto = request.args.get('idobjeto')
    archivo = request.args.get('archivo')

    pathx=os.path.join(app.root_path, 'static', 'uploads',tablename,idobjeto,archivo)

    wb = openpyxl.load_workbook(pathx)

    sheet = wb.active

    #x1 = sheet['A1']
    #x2 = sheet['A2']
    #using cell() function
    #x3 = sheet.cell(row=3, column=1)

    #parseando filas

    tienecontenido=''
    fila=0
    columna=0

    while tienecontenido!='':
        tienecontenido=sheet.cell(row=fila,columna=0)

        tienecontenido2=''
        campos=''
        while tienecontenido2!='':
            xfila=sheet.cell(row=fila,column=columna)
            columna+=1
            campos+="-" + xfila
            print(campos)

        fila+=1


    return 'ok'



@app.route('/agendaget', methods=['GET','POST'])
def agendaget():

    campofiltro = request.args.get('campofiltro')
    idobjetofiltro = request.args.get('idobjetofiltro')




    return agendadatos(campofiltro,idobjetofiltro)



@app.route('/tblarticulo')
def tblarticulo():
    return render_template('tables/tblarticulo_form.html',topmenu=topmenutemplate())

@app.route('/backup', methods=['GET','POST'])
def backup():

    id = request.args.get('id')
    tablename = request.args.get('tablename')


    if path.exists((os.path.join(app.root_path, 'static', 'uploads',tablename)))==False :
                   #creando el folder tabla
                   os.mkdir(os.path.join(app.root_path, 'static', 'uploads',tablename))

    if path.exists((os.path.join(app.root_path, 'static', 'uploads',tablename,id)))==False :
                   os.mkdir(os.path.join(app.root_path, 'static', 'uploads',tablename,id))

    pathx=os.path.join(app.root_path, 'static', 'uploads',tablename,id)

    filestamp = time.strftime('%Y-%m-%d-%I:%M')

    elcomando="mysqldump -u " + userx + " -p" + passwdx + " -h " + hostx + " " + databasex + " > " + pathx + '/' + databasex + "_"+filestamp + ".sql"
    #userx,passwdx,hostx,databasex,databasex+"_"+filestamp
    os.popen(elcomando)

    return elcomando

@app.route('/restaurarbackup', methods=['GET','POST'])
def restaurarbackup():

    id = request.args.get('id')
    tablename = request.args.get('tablename')
    name = request.args.get('name')


    rutafile=os.path.join(app.root_path, 'static', 'uploads',tablename,id,name)


    elcomando="mysql -u " + userx + " -p" + passwdx + " -h " + hostx + " " + databasex + " < " + rutafile
    #userx,passwdx,hostx,databasex,databasex+"_"+filestamp
    os.popen(elcomando)

    return elcomando

@app.route('/excelexport', methods=['GET','POST'])
def excelexport():

    buscar = request.args.get('buscar')
    tablename = request.args.get('tablename')
    CAMPODISPLAY_startsWith = request.args.get('CAMPODISPLAY_startsWith')
    campofiltro = request.args.get('campofiltro')
    idobjetofiltro = request.args.get('idobjetofiltro')

    data=''




    if buscar=='' or buscar=='null' or buscar==None:
        data=populatexlsx(tablename,CAMPODISPLAY_startsWith,campofiltro,idobjetofiltro)

    else:
        data=buscadorxlsx("select *,concat('<a  style=""color:blue"" href=""javascript:modoedicion(',ID" + tablename.replace('tbl','') + ",');"">Editar</a>') as EDITAR from " + tablename.lower() + " where " + quierolaclasusulawhere(tablename,buscar))




    pathx=os.path.join(app.root_path, 'static', 'xlsx','modelo.xlsx')
    wb = load_workbook(pathx)
    sheet = wb.active

    #data = (
    #(11, 48, 50),
    #(81, 30, 82),
    #(20, 51, 72),
    #(21, 14, 60),
    #(28, 41, 49),
    #(74, 65, 53),
    #("Peter", 'Andrew',45.63)
#)



    for i in data:
        sheet.append(i)

    elarchivo="Excel_Export" + str(randint(999,999999)) + ".xlsx"

    pathx2=os.path.join(app.root_path, 'static', 'temp',elarchivo)
    wb.save(pathx2)

    ruta='/static/temp/' + elarchivo

    return render_template("resultado_exportacion.html",ruta=ruta)

@app.route('/validapropiedad', methods=['GET','POST'])
def validapropiedad():

    idobjeto = request.args.get('idobjeto')
    esadmin = request.args.get('esadmin')
    tablename = request.args.get('tablename')
    idusername = request.args.get('idusername')
    tipouser = request.args.get('tipouser')


    #print(populatepropiedad(tablename,idusername,tipouser,esadmin,idobjeto) + 'ddddddddddddddddddddddddddddddddddddddddddddd')


    return populatepropiedad(tablename,idusername,tipouser,esadmin,idobjeto)



@app.route('/populator', methods=['GET','POST'])
def populator():

    esadmin = request.args.get('esadmin')
    buscar = request.args.get('buscar')
    tablename = request.args.get('tablename')
    CAMPODISPLAY_startsWith = request.args.get('CAMPODISPLAY_startsWith')
    campofiltro = request.args.get('campofiltro')
    idobjetofiltro = request.args.get('idobjetofiltro')
    usuario = request.args.get('usuario')
    idusername = request.args.get('idusername')
    tipouser = request.args.get('tipouser')


    if CAMPODISPLAY_startsWith!='+':
        #REGISTRO DE PRODUCTIVIDAD
        productividad("VISITA PÁGINA",idusername,usuario,tablename)


    if buscar=='' or buscar=='null' or buscar==None:
        return populate(tablename,CAMPODISPLAY_startsWith,campofiltro,idobjetofiltro,usuario,idusername,tipouser,esadmin)
    else:
        return buscador("select *,concat('<a style=""color:blue"" href=""javascript:modoedicion(',ID" + tablename.replace('tbl','') + ",');"">Editar</a>') as EDITAR from " + tablename.lower() + " where " + quierolaclasusulawhere(tablename,buscar),usuario,idusername,tipouser,esadmin)


@app.route('/displayquery', methods=['GET','POST'])
def displayquery():
    query = request.args.get('query')
    return make_response(populatequery(query), 200)

@app.route('/displayquery2', methods=['GET','POST'])
def displayquery2():

    query = request.args.get('query')
    cwhere = request.args.get('where')


    #print ("el tipo es " + str(type(cwhere) ))

    if str(type(cwhere))!="<class 'NoneType'>"  and  str(cwhere)!='' and  str(cwhere)!=' ' :
        cwhere=' where ' + str(cwhere)
    else:
        cwhere=''




    #if cwhere=='None' or cwhere=='' or cwhere=='null':
    #    cwhere=''

    return make_response(populatequery2(query,cwhere), 200)



@app.route('/numregistros', methods=['GET','POST'])
def numregistros():

    tablename = request.args.get('tabla')


    return cuentaregistros(tablename)

@app.route('/pivotstate', methods=['GET','PUT'])
def pivotstate():


    metodo = request.args.get('tablename')

    print(metodo)

    return 'ok'

@app.route('/dameultimoid', methods=['GET','POST'])
def dameultimoidx():

    tablename = request.args.get('tablename')


    return dameultimoid(tablename)


@app.route('/dametablasbuild', methods=['GET','POST'])
def dametablasbuild():

    return tablasparabuild()



@app.route('/cacheregen', methods=['GET','POST'])
def cacheregenx():

    cacheregen()

    return 'ok'

@app.route('/validaracceso', methods=['GET','POST'])
def validaracceso():

    idusername = request.args.get('idusername')
    tablename = request.args.get('tablename')

    #print(checarpermisos(tablename,idusername) + 'HHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHH')


    return str(checarpermisos(tablename,idusername))


@app.route('/purgar', methods=['GET','POST'])
def purgarX():

    tablename = request.args.get('tablename')

    purgar(tablename)

    return 'Elementos eliminados'



######################################################################################################
######################################################################################################


@app.route('/importadorxlsx', methods=['GET','POST'])
def importadorxlsx():

    tablename = request.args.get('tablename')

    return render_template("tables/" + tablename + "_importador.html")


 

